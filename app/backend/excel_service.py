"""Encrypted Excel data access (cross-platform).

The inventory lives in a password-protected ``.xlsx``. We use ``msoffcrypto`` to
decrypt for reading and to re-encrypt after writing, with ``openpyxl`` handling
the spreadsheet content in between. This is pure Python, so it behaves identically
on Ubuntu (dev) and Windows (prod) and needs **no** Microsoft Excel install.

The canonical sheet is ``همه`` (all items). ``total_qty`` is a formula in the
source file (``carton_count × qty_per_carton``); we read its cached value and, if
absent, compute it. On write we store literal integers so the headless app stays
self-consistent.

Roles: the *modify* (read-write) password is enforced at the application layer —
only a writable service (created for write-capable roles) performs saves.
"""

from __future__ import annotations

import io
import logging
from pathlib import Path

import msoffcrypto
import openpyxl
from msoffcrypto.format.ooxml import OOXMLFile

from app.backend.models import INVENTORY_FIELDS, Product, _coerce

logger = logging.getLogger("hyper_samen.excel")

CANONICAL_SHEET = "همه"
_HEADER_HINTS = {"نام", "بارکد", "نوع", "مدل", "شرکت", "تعداد"}
_N_COLS = len(INVENTORY_FIELDS)


class ExcelError(Exception):
    """Raised for any inventory-file access problem (wrong password, missing file…)."""


def _normalize(text) -> str:
    if text is None:
        return ""
    return " ".join(str(text).split())


def _looks_like_header(row: tuple) -> bool:
    cells = {_normalize(c) for c in row if c is not None}
    hits = sum(1 for hint in _HEADER_HINTS for c in cells if hint in c)
    return hits >= 3


class ExcelService:
    """Read/write the encrypted inventory workbook."""

    def __init__(self, path: Path, read_password: str, write_password: str = "",
                 writable: bool = False):
        self._path = Path(path)
        self._read_pw = read_password
        self._write_pw = write_password
        self._writable = writable

    @property
    def writable(self) -> bool:
        return self._writable

    @property
    def backend_name(self) -> str:
        return "crypto"

    # -- decryption / encryption ---------------------------------------------

    def _decrypt_bytes(self) -> io.BytesIO:
        if not self._path.exists():
            raise ExcelError(f"فایل موجودی یافت نشد: {self._path}")
        try:
            with open(self._path, "rb") as f:
                office = msoffcrypto.OfficeFile(f)
                office.load_key(password=self._read_pw)
                buf = io.BytesIO()
                office.decrypt(buf)
        except Exception as exc:  # noqa: BLE001 - surface as a domain error
            raise ExcelError("رمز فایل اکسل نادرست است یا فایل خراب است.") from exc
        buf.seek(0)
        return buf

    def _load_workbook(self, data_only: bool):
        buf = self._decrypt_bytes()
        return openpyxl.load_workbook(buf, data_only=data_only)

    def _encrypt_and_save(self, plain: io.BytesIO) -> None:
        plain.seek(0)
        enc = io.BytesIO()
        OOXMLFile(plain).encrypt(self._read_pw, enc)
        enc.seek(0)
        tmp = self._path.with_suffix(self._path.suffix + ".tmp")
        with open(tmp, "wb") as f:
            f.write(enc.read())
        tmp.replace(self._path)

    # -- helpers --------------------------------------------------------------

    def _sheet(self, wb):
        if CANONICAL_SHEET in wb.sheetnames:
            return wb[CANONICAL_SHEET]
        return wb.active

    def _header_row_index(self, ws) -> int:
        """1-based index of the header row (defaults to 1)."""
        for idx, row in enumerate(ws.iter_rows(min_row=1, max_row=6, values_only=True), start=1):
            if _looks_like_header(row):
                return idx
        return 1

    # -- public API -----------------------------------------------------------

    def _parse_sheet(self, ws) -> list[Product]:
        """Read products from a worksheet (canonical column order, after the header)."""
        header_idx = self._header_row_index(ws)
        products: list[Product] = []
        for r, row in enumerate(
            ws.iter_rows(min_row=header_idx + 1, values_only=True), start=header_idx + 1
        ):
            cells = list(row[:_N_COLS]) + [None] * (_N_COLS - len(row))
            if all(c is None or (isinstance(c, str) and not c.strip()) for c in cells):
                continue
            product = Product.from_cells(cells, row=r)
            if product.total_qty is None and product.carton_count and product.qty_per_carton:
                product.total_qty = product.carton_count * product.qty_per_carton
            if not product.name:
                continue
            products.append(product)
        return products

    def read_products(self) -> list[Product]:
        wb = self._load_workbook(data_only=True)
        ws = self._sheet(wb)
        products = self._parse_sheet(ws)
        logger.info("Loaded inventory: %d products from sheet '%s'", len(products), ws.title)
        return products

    def read_external(self, path: Path, password: str | None = None) -> list[Product]:
        """Parse products from an external .xlsx (plain, or encrypted with *password*)."""
        path = Path(path)
        if not path.exists():
            raise ExcelError(f"فایل ورودی یافت نشد: {path}")
        with open(path, "rb") as f:
            try:
                office = msoffcrypto.OfficeFile(f)
                is_encrypted = office.is_encrypted()
            except Exception:  # noqa: BLE001 - not an encrypted OLE container → plain xlsx
                is_encrypted = False
            if is_encrypted:
                try:
                    office.load_key(password=password or "")
                    buf = io.BytesIO()
                    office.decrypt(buf)
                except Exception as exc:  # noqa: BLE001
                    raise ExcelError("فایل ورودی رمزدار است؛ رمز واردشده نادرست است.") from exc
                buf.seek(0)
                wb = openpyxl.load_workbook(buf, data_only=True)
            else:
                try:
                    wb = openpyxl.load_workbook(path, data_only=True)
                except Exception as exc:  # noqa: BLE001
                    raise ExcelError("فایل ورودی معتبر نیست یا قابل خواندن نیست.") from exc
        ws = self._sheet(wb)
        products = self._parse_sheet(ws)
        logger.info("Read %d products from external file '%s'", len(products), path.name)
        return products

    def replace_all(self, products: list[Product]) -> int:
        """Replace every data row of the canonical sheet with *products*."""
        self._require_writable()
        wb = self._load_workbook(data_only=False)
        ws = self._sheet(wb)
        header_idx = self._header_row_index(ws)
        if ws.max_row > header_idx:
            ws.delete_rows(header_idx + 1, ws.max_row - header_idx)
        for i, p in enumerate(products, start=header_idx + 1):
            p.row = i
            self._write_row(ws, p)
        plain = io.BytesIO(); wb.save(plain)
        self._encrypt_and_save(plain)
        logger.info("Replaced inventory with %d products", len(products))
        return len(products)

    def bulk_upsert(self, products: list[Product]) -> tuple[int, int]:
        """Update rows matching by barcode, append the rest. Returns (updated, added)."""
        self._require_writable()
        wb = self._load_workbook(data_only=False)
        ws = self._sheet(wb)
        header_idx = self._header_row_index(ws)
        bc_col = INVENTORY_FIELDS.index("barcode") + 1
        existing: dict[str, int] = {}
        for r in range(header_idx + 1, ws.max_row + 1):
            bc = _coerce("barcode", ws.cell(row=r, column=bc_col).value)
            if bc:
                existing[bc] = r
        next_row = ws.max_row + 1
        updated = added = 0
        for p in products:
            if p.barcode and p.barcode in existing:
                p.row = existing[p.barcode]
                self._write_row(ws, p)
                updated += 1
            else:
                p.row = next_row
                self._write_row(ws, p)
                if p.barcode:
                    existing[p.barcode] = next_row
                next_row += 1
                added += 1
        plain = io.BytesIO(); wb.save(plain)
        self._encrypt_and_save(plain)
        logger.info("Bulk upsert: %d updated, %d added", updated, added)
        return updated, added

    def _require_writable(self) -> None:
        if not self._writable:
            raise ExcelError("شما اجازه ویرایش موجودی را ندارید.")

    def _write_row(self, ws, product: Product) -> None:
        for col, fld in enumerate(INVENTORY_FIELDS, start=1):
            ws.cell(row=product.row, column=col).value = getattr(product, fld)

    def update_product(self, product: Product) -> None:
        self._require_writable()
        if not product.row:
            raise ExcelError("ردیف کالا برای به‌روزرسانی مشخص نیست.")
        wb = self._load_workbook(data_only=False)
        ws = self._sheet(wb)
        self._write_row(ws, product)
        plain = io.BytesIO(); wb.save(plain)
        self._encrypt_and_save(plain)
        logger.info("Updated product at row %d: %s", product.row, product.name)

    def append_product(self, product: Product) -> int:
        self._require_writable()
        wb = self._load_workbook(data_only=False)
        ws = self._sheet(wb)
        new_row = ws.max_row + 1
        product.row = new_row
        self._write_row(ws, product)
        plain = io.BytesIO(); wb.save(plain)
        self._encrypt_and_save(plain)
        logger.info("Added new product at row %d: %s", new_row, product.name)
        return new_row

    def delete_product(self, product: Product) -> None:
        self._require_writable()
        if not product.row:
            raise ExcelError("ردیف کالا برای حذف مشخص نیست.")
        wb = self._load_workbook(data_only=False)
        ws = self._sheet(wb)
        ws.delete_rows(product.row, 1)
        plain = io.BytesIO(); wb.save(plain)
        self._encrypt_and_save(plain)
        logger.info("Deleted product at row %d: %s", product.row, product.name)


def encrypt_workbook(wb, path: Path, password: str) -> None:
    """Save an openpyxl workbook to *path*, password-encrypted (msoffcrypto)."""
    plain = io.BytesIO()
    wb.save(plain)
    plain.seek(0)
    enc = io.BytesIO()
    OOXMLFile(plain).encrypt(password, enc)
    enc.seek(0)
    path = Path(path)
    path.parent.mkdir(parents=True, exist_ok=True)
    with open(path, "wb") as f:
        f.write(enc.read())
    logger.info("Encrypted workbook written: %s", path)


def create_excel_service(config, writable: bool) -> ExcelService:
    """Factory: build an :class:`ExcelService` from a :class:`Config`."""
    return ExcelService(
        path=config.inventory_file,
        read_password=config.excel_read_password,
        write_password=config.excel_write_password,
        writable=writable,
    )
